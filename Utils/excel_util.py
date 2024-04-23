#excel工具类，负责创建excel文件并写入实验结果数据
import os

import pandas as pd
import openpyxl


def create_workbook(filename):
    if os.path.exists(filename):
        # 如果文件存在，打开现有的工作簿
        workbook = openpyxl.load_workbook(filename)
    else:
        # 如果文件不存在，创建新的工作簿
        workbook = openpyxl.Workbook()
    print(workbook.sheetnames)
    return workbook


def create_sheet(sheet_name, wb):
    if "Sheet" in wb.sheetnames:  # 初始列表
        sheet = wb.active
        sheet.title = sheet_name
    elif sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])
        sheet = wb.create_sheet(title=sheet_name)
    else:
        sheet = wb.create_sheet(title=sheet_name)
    return sheet


def init_table_head(sheet, headers):
    for index, header in enumerate(headers):
        sheet.cell(row=1, column=index + 1).value = header
    return sheet


def append_row(sheet, row):
    sheet.append(row)
    return sheet


def calculate_average(sheet):
    # 获取表格的最大行和列
    max_row = sheet.max_row
    max_column = sheet.max_column

    # 计算每一列的平均值
    for col_num in range(2, max_column + 1):
        # 跳过标题行，从第2行开始计算平均值
        values = [float(sheet.cell(row=row, column=col_num).value) for row in range(2, max_row + 1)]
        average = sum(values) / len(values) if values else 0

        # 在最后一行插入平均值
        sheet.cell(row=max_row + 1, column=col_num, value=average)
    sheet.cell(row=max_row + 1, column=1, value="平均数")

if __name__ == "__main__":
    create_workbook("sds")
